#!/usr/bin/env python3
"""
Import Demo Log Excel file into trial management system
Reads "Request for trial licenses" sheet and creates organizations and trial users
"""

import os
import sys
import json
from datetime import datetime
from pathlib import Path
import re

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("Error: supabase-py not installed. Install with: pip3 install supabase")
    sys.exit(1)

# Configuration
EXCEL_FILE = "/Users/sriharsha/Desktop/myRA AI Demo Log.xlsx"
SHEET_NAME = "Request for trial licenses"

class DemoLogImporter:
    def __init__(self):
        # Initialize Supabase client
        supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

        if not supabase_url or not supabase_key:
            raise ValueError("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY environment variables")

        self.supabase = create_client(supabase_url, supabase_key)
        self.import_log = []
        self.errors = []

    def log(self, message):
        """Log import messages"""
        print(message)
        self.import_log.append(message)

    def error(self, message):
        """Log errors"""
        print(f"ERROR: {message}")
        self.errors.append(message)

    def parse_date(self, date_str):
        """
        Parse various date formats
        - "4th October" or "4th October 2024"
        - "6th October" or "6th October 2024"
        - None or empty string returns None
        """
        if not date_str or date_str.lower() == 'none':
            return None

        try:
            date_str = str(date_str).strip()

            # Try parsing with year (if provided)
            for fmt in ["%d%B %Y", "%d %B %Y", "%d %b %Y", "%d%b %Y"]:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.isoformat()
                except ValueError:
                    continue

            # Try parsing without year (assume current year)
            for fmt in ["%d%B", "%d %B", "%d %b", "%d%b"]:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    # Assume current year
                    dt = dt.replace(year=datetime.now().year)
                    return dt.isoformat()
                except ValueError:
                    continue

            # If we get here, log warning but don't fail
            self.log(f"  ⚠️  Could not parse date format: '{date_str}'. Using None.")
            return None

        except Exception as e:
            self.log(f"  ⚠️  Date parsing error for '{date_str}': {str(e)}")
            return None

    def validate_email(self, email):
        """Basic email validation"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, str(email)) is not None

    def read_excel_data(self):
        """Read and parse Excel file"""
        self.log(f"\n📖 Reading Excel file: {EXCEL_FILE}")

        if not os.path.exists(EXCEL_FILE):
            raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

        wb = openpyxl.load_workbook(EXCEL_FILE)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{SHEET_NAME}' not found in workbook. Available sheets: {wb.sheetnames}")

        ws = wb[SHEET_NAME]
        self.log(f"✅ Found sheet: '{SHEET_NAME}'")

        # Parse headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        self.log(f"📋 Columns: {headers}")

        # Parse data rows
        data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_num, cell in enumerate(row):
                header = headers[col_num] if col_num < len(headers) else f"Column_{col_num}"
                row_data[header] = cell.value

            # Skip empty rows
            if any(v is not None for v in row_data.values()):
                data.append(row_data)

        self.log(f"✅ Read {len(data)} data rows")
        return data

    def delete_existing_orgs(self):
        """Delete all existing trial organizations (CASCADE will handle related records)"""
        self.log("\n🗑️  Deleting existing trial organizations...")

        try:
            # Get count before deletion
            result = self.supabase.table('trial_organizations').select('*').execute()
            count = len(result.data) if result.data else 0

            if count > 0:
                self.log(f"   Found {count} existing organizations")

                # Delete all organizations (CASCADE will handle related records)
                response = self.supabase.table('trial_organizations').delete().neq('org_id', 'null').execute()
                self.log(f"✅ Deleted {count} organizations and all related records")
            else:
                self.log("   No existing organizations found")

        except Exception as e:
            raise Exception(f"Failed to delete existing organizations: {str(e)}")

    def create_organization(self, company_data):
        """Create a trial organization"""
        try:
            org_name = company_data.get('Company Name', '').strip()
            domain = company_data.get('Domain', '').strip()
            sales_poc = company_data.get('Sales POC', '').strip()
            comments = company_data.get('Comments') or None

            if not org_name:
                raise ValueError("Company Name is required")

            # Prepare organization data
            org_data = {
                'org_name': org_name,
                'org_domain': domain if domain and domain.lower() != 'none' else None,
                'sales_poc': sales_poc if sales_poc and sales_poc.lower() != 'none' else None,
                'status': 'active',  # Smart default
                'comments': comments,
            }

            # Create organization
            response = self.supabase.table('trial_organizations').insert(org_data).execute()

            if not response.data or len(response.data) == 0:
                raise Exception("Failed to create organization")

            org = response.data[0]
            self.log(f"  ✅ Created organization: {org_name} (ID: {org['org_id']})")
            return org

        except Exception as e:
            raise Exception(f"Failed to create organization '{org_name}': {str(e)}")

    def create_trial_user(self, org_id, user_data):
        """Create a trial user"""
        try:
            email = user_data.get('Email', '').strip()
            designation = user_data.get('Title/Role', '').strip()
            license_date = user_data.get('License given on')

            if not email:
                raise ValueError("Email is required")

            if not self.validate_email(email):
                raise ValueError(f"Invalid email format: {email}")

            # Determine user status based on license date
            parsed_date = self.parse_date(license_date) if license_date else None
            user_status = 'active' if parsed_date else 'invited'

            # Prepare user data
            user_data_insert = {
                'org_id': org_id,
                'email': email,
                'user_status': user_status,
                'designation': designation if designation else None,
                'access_enabled_at': parsed_date,
            }

            # Create trial user
            response = self.supabase.table('trial_users').insert(user_data_insert).execute()

            if not response.data or len(response.data) == 0:
                raise Exception("Failed to create trial user")

            user = response.data[0]
            self.log(f"  ✅ Created trial user: {email} (Status: {user_status})")
            return user

        except Exception as e:
            raise Exception(f"Failed to create trial user: {str(e)}")

    def import_data(self):
        """Main import process"""
        try:
            self.log("\n" + "="*60)
            self.log("MyRA Demo Log Import - Trial Management System")
            self.log("="*60)

            # Step 1: Read Excel data
            excel_data = self.read_excel_data()

            # Step 2: Delete existing organizations
            self.delete_existing_orgs()

            # Step 3: Import data
            self.log("\n📝 Importing data...")

            for idx, row_data in enumerate(excel_data, start=1):
                self.log(f"\n--- Processing Row {idx} ---")

                try:
                    company_name = row_data.get('Company Name', '').strip()
                    email = row_data.get('Email', '').strip()

                    self.log(f"   Company: {company_name}")
                    self.log(f"   Email: {email}")

                    # Create organization
                    org = self.create_organization(row_data)

                    # Create trial user
                    if email:
                        user = self.create_trial_user(org['org_id'], row_data)
                    else:
                        self.log(f"  ⚠️  No email provided, skipping trial user creation")

                except Exception as e:
                    self.error(f"Row {idx}: {str(e)}")

            # Step 4: Summary
            self.print_summary()

        except Exception as e:
            self.error(f"Import failed: {str(e)}")
            return False

        return len(self.errors) == 0

    def print_summary(self):
        """Print import summary"""
        self.log("\n" + "="*60)
        self.log("IMPORT SUMMARY")
        self.log("="*60)

        # Get final counts
        try:
            orgs = self.supabase.table('trial_organizations').select('*').execute()
            users = self.supabase.table('trial_users').select('*').execute()

            org_count = len(orgs.data) if orgs.data else 0
            user_count = len(users.data) if users.data else 0

            self.log(f"\n✅ Import completed successfully!")
            self.log(f"   Organizations created: {org_count}")
            self.log(f"   Trial users created: {user_count}")

            if self.errors:
                self.log(f"\n⚠️  Errors encountered: {len(self.errors)}")
                for error in self.errors:
                    self.log(f"   - {error}")
            else:
                self.log("\n✅ No errors encountered")

        except Exception as e:
            self.log(f"Error getting summary: {str(e)}")

        self.log("\n" + "="*60)

def main():
    """Main entry point"""
    try:
        importer = DemoLogImporter()
        success = importer.import_data()

        if not success:
            sys.exit(1)

    except Exception as e:
        print(f"FATAL ERROR: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
